from .serializers import ClienteSerializer, ProveedorSerializer, FacturaSerializer, UserSerializer, FacturaCreateUpdateSerializer
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from django_filters.rest_framework import DjangoFilterBackend
from .models import Cliente, Proveedor, Factura, User
from rest_framework import viewsets,generics, status
from rest_framework.filters import SearchFilter
from .models import Proveedor, FacturaProveedor, Factura
from .permissions import IsAdmin, IsContador
from rest_framework.response import Response
from django.db.models import Sum, Count, Q
from rest_framework.views import APIView
from django.utils.timezone import now
from .services import obtener_alertas_facturas
from .serializers import *
from django.http import HttpResponse
import csv, openpyxl
from openpyxl.styles import Font
from django.db import transaction
from rest_framework.parsers import MultiPartParser, FormParser
class DashboardDataView(APIView):
    def get(self, request):
        # Calcular flujo de caja (por cobrar y por pagar) agrupado por mes
        flujo_caja = Factura.objects.annotate(
            mes=F('fecha_emision__month')  # Extraemos el mes
        ).values('mes').annotate(
            total_por_cobrar=Sum('monto_total', filter=Q(estado='Pendiente')),
            total_por_pagar=Sum('monto_total', filter=Q(estado='Pagada'))
        ).order_by('mes')

        # Contar facturas por estado (para el gráfico de pastel)
        estado_facturas = Factura.objects.values('estado').annotate(
            cantidad=Count('id')
        )

        # Formatear los datos para el frontend
        data = {
            'flujo_caja': [
                {
                    'mes': mes['mes'],
                    'nombre_mes': now().replace(month=mes['mes']).strftime('%B'),
                    'total_por_cobrar': mes['total_por_cobrar'] or 0,
                    'total_por_pagar': mes['total_por_pagar'] or 0,
                }
                for mes in flujo_caja
            ],
            'estado_facturas': list(estado_facturas),
        }

        return Response(data)

class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer

class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer

class FacturaViewSet(viewsets.ModelViewSet):
    queryset = Factura.objects.all()
    serializer_class = FacturaSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['cliente', 'estado', 'fecha_emision', 'fecha_vencimiento']
    search_fields = ['numero_factura']

    def get_serializer_class(self):
        # Usa diferentes serializers para listar y crear/actualizar
        if self.action in ['create', 'update', 'partial_update']:
            return FacturaCreateUpdateSerializer
        return FacturaSerializer

    def create(self, request, *args, **kwargs):
        # Actualiza el estado automáticamente al crear
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        factura = serializer.save()
        factura.actualizar_estado()
        return Response(FacturaSerializer(factura).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        # Actualiza el estado automáticamente al editar
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        factura = serializer.save()
        factura.actualizar_estado()
        return Response(FacturaSerializer(factura).data)

class LoginView(TokenObtainPairView):
    permission_classes = [AllowAny]

class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
            "message": "Usuario creado exitosamente",
            "user": {
                "username": user.username,
                "email": user.email,
                "role": user.role,
            }
        }, status=status.HTTP_201_CREATED)
    
class AdminView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        return Response({"message": "Solo los administradores pueden ver esto."})

class ContadorView(APIView):
    permission_classes = [IsContador]

    def get(self, request):
        return Response({"message": "Solo los contadores pueden ver esto."})
    
class AdminOnlyView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        return Response({"message": "Acceso permitido: eres Administrador"})
    
class FacturaProveedorViewSet(viewsets.ModelViewSet):
    queryset = FacturaProveedor.objects.all()
    serializer_class = FacturaProveedorSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['proveedor', 'estado', 'fecha_emision', 'fecha_vencimiento']
    search_fields = ['numero_factura']

    def get_serializer_class(self):
        # Usa diferentes serializers para listar y crear/actualizar
        if self.action in ['create', 'update', 'partial_update']:
            return FacturaProveedorCreateUpdateSerializer
        return FacturaProveedorSerializer

    def create(self, request, *args, **kwargs):
        # Actualiza el estado automáticamente al crear
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        factura = serializer.save()
        factura.actualizar_estado()
        return Response(FacturaProveedorSerializer(factura).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        # Actualiza el estado automáticamente al editar
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        factura = serializer.save()
        factura.actualizar_estado()
        return Response(FacturaProveedorSerializer(factura).data)
    
class AlertasFacturasView(APIView):
    def get(self, request):
        alertas = obtener_alertas_facturas()
        proximas_vencer = FacturaSerializer(alertas['proximas_vencer'], many=True).data
        vencidas = FacturaSerializer(alertas['vencidas'], many=True).data

        return Response({
            "proximas_vencer": proximas_vencer,
            "vencidas": vencidas,
        })

class ExportarFacturasCSVView(APIView):
    def get(self, request):
        # Crear la respuesta con el encabezado CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="facturas.csv"'

        # Escribir los encabezados del CSV
        writer = csv.writer(response)
        writer.writerow(['Número de Factura', 'Cliente', 'Fecha de Emisión', 'Fecha de Vencimiento', 'Monto Total', 'Estado'])

        # Escribir los datos de las facturas
        facturas = Factura.objects.all()
        for factura in facturas:
            writer.writerow([
                factura.numero_factura,
                factura.cliente.nombre,
                factura.fecha_emision,
                factura.fecha_vencimiento,
                factura.monto_total,
                factura.estado,
            ])

        return response
    
class ExportarFacturasExcelView(APIView):
    def get(self, request):
        # Crear un archivo Excel en memoria
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Facturas'

        # Encabezados
        headers = ['Número de Factura', 'Cliente', 'Fecha de Emisión', 'Fecha de Vencimiento', 'Monto Total', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Escribir los datos
        facturas = Factura.objects.all()
        for row_num, factura in enumerate(facturas, 2):
            ws.cell(row=row_num, column=1, value=factura.numero_factura)
            ws.cell(row=row_num, column=2, value=factura.cliente.nombre)
            ws.cell(row=row_num, column=3, value=factura.fecha_emision.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=4, value=factura.fecha_vencimiento.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=5, value=factura.monto_total)
            ws.cell(row=row_num, column=6, value=factura.estado)

        # Crear la respuesta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="facturas.xlsx"'

        # Guardar el archivo en la respuesta
        wb.save(response)

        return response
    
class ImportarFacturasCSVView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            return Response({'error': 'No se proporcionó un archivo.'}, status=400)

        try:
            with transaction.atomic():
                csv_file = csv.reader(file.read().decode('utf-8').splitlines())
                next(csv_file)  # Saltar la fila de encabezados

                facturas_omitidas = []  # Para registrar facturas duplicadas

                for row in csv_file:
                    numero_factura, cliente_nombre, fecha_emision, fecha_vencimiento, monto_total, estado = row

                    # Verificar si el número de factura ya existe
                    if Factura.objects.filter(numero_factura=numero_factura).exists():
                        facturas_omitidas.append(numero_factura)
                        continue  # Ignorar facturas duplicadas

                    # Crear cliente si no existe
                    cliente, _ = Cliente.objects.get_or_create(nombre=cliente_nombre)

                    # Crear factura
                    Factura.objects.create(
                        numero_factura=numero_factura,
                        cliente=cliente,
                        fecha_emision=fecha_emision,
                        fecha_vencimiento=fecha_vencimiento,
                        monto_total=monto_total,
                        estado=estado,
                    )

            if facturas_omitidas:
                return Response({
                    'message': 'Facturas importadas con éxito.',
                    'facturas_omitidas': facturas_omitidas,
                }, status=201)
            else:
                return Response({'message': 'Todas las facturas fueron importadas correctamente.'}, status=201)

        except Exception as e:
            return Response({'error': str(e)}, status=400)
